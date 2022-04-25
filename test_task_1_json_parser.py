"""
Parse a json files in .input/ folder and write them to separate sheets
in Excel book for each json file. Output file will be in the same folder
as script and named "task1.xlsx"
"""

import os
import json

from openpyxl import Workbook

# const:
_folder = "input"
_headers_key = "headers"
_values_key = "values"
_props = "properties"
_table_header = "QuickInfo"
_column = "X"
_row = "Y"
_table_content = "Text"


def get_files_list(folder: str, files: list[str] = []) -> list[str]:
    '''Create list with paths to files in folder .input/'''
    for filename in os.listdir(_folder):
        filepath = os.path.join(_folder, filename)
        files.append(filepath)
    return sorted(files)


def load_file(filename: str) -> dict:
    '''Open json file from previously generated path to it. Returns a dict.'''
    with open(filename) as f:
        file_content = f.read()
        templates = json.loads(file_content)
        return templates


def get_table_info(
        property_type: str,
        templates: dict,
        key_type: str,
        props: str,
) -> list[str]:
    '''
    Used to get various info for writing in cells: table headers,
    headers id (for keeping right ordering in cells) and rows id for future table.
    The <property_type> argument is used to select which kind of info needs to return.
    Output is a list.
    '''
    table_info: list[str] = []
    for item in templates[key_type]:
        if item[props][property_type] not in table_info:
            table_info.append(item[props][property_type])
    return table_info


def get_table_data(
        table_rows_list: list[str],
        table_columns_list: list[str],
        values_key: str,
        props: str,
        row: str,
        column: str,
        table_content: str,
        record: list[str] = [],
        data: dict = {},
) -> dict:
    '''
    Collects data from previously opened json for writing in table cells.
    Keeps ordering and returns a dict in following format: {<id>: <record>}
    '''
    for table_row in table_rows_list:
        for table_column in table_columns_list:
            for item in templates[values_key]:
                if (
                        item[props][row] == table_row
                        and item[props][column] == table_column
                ):
                    record.append(item[props][table_content])
        data[table_row] = tuple(record)
        record.clear()
    return data


files_list = get_files_list(_folder)

wb = Workbook()
excel_file = 'task1.xlsx'

sheets = ['sheet' + str(i) for i, _ in enumerate(files_list)]

for i, filename in enumerate(files_list):
    templates: dict = load_file(filename)
    table_headers_list: list[str] = get_table_info(_table_header, templates, _headers_key, _props)
    table_columns_list: list[str] = get_table_info(_column, templates, _headers_key, _props)
    table_rows_list: list[str] = get_table_info(_row, templates, _values_key, _props)

    table_data: dict = get_table_data(
        table_rows_list,
        table_columns_list,
        _values_key,
        _props,
        _row,
        _column,
        _table_content
    )

    sheets[i] = wb.create_sheet(title=f"Sheet{i}")
    ws = wb[f"Sheet{i}"]
    title_row = ws.row_dimensions[1]
    sheets[i].append(table_headers_list)
    for key, value in table_data.items():
        sheets[i].append(value)
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

wb.save(filename=excel_file)

